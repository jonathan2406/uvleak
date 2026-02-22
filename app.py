"""
InternLink - Sistema de Gestión de Pasantías Universitarias
Plataforma para conectar estudiantes, empresas y coordinadores académicos
"""
from flask import (
    Flask, render_template, request, redirect, url_for,
    jsonify, make_response, send_file, abort
)
from functools import wraps
import json
import hashlib
import re
import secrets
import os
from datetime import datetime
from urllib.request import urlopen, Request
from urllib.parse import quote
from urllib.error import URLError
import jwt
from werkzeug.utils import secure_filename
from io import BytesIO
from dotenv import load_dotenv
from upstash_redis import Redis
import cloudinary
import cloudinary.uploader
import cloudinary.api
from cloudinary import utils as cloudinary_utils

load_dotenv()

# Configurar Cloudinary: priorizar CLOUDINARY_URL, luego variables individuales
# NOTA: Esta configuración se ejecuta al importar el módulo, lo cual es correcto para serverless
cloudinary_url_env = os.getenv('CLOUDINARY_URL')
cloudinary_configured = False
try:
    if cloudinary_url_env:
        # Si está CLOUDINARY_URL, Cloudinary la lee automáticamente del entorno
        # cloudinary.config() sin parámetros lee CLOUDINARY_URL automáticamente
        cloudinary.config()
        cloudinary_configured = True
    else:
        # Fallback a variables individuales
        cloud_name = os.getenv('CLOUDINARY_CLOUD_NAME', '')
        api_key = os.getenv('CLOUDINARY_API_KEY', '')
        api_secret = os.getenv('CLOUDINARY_API_SECRET', '')
        if cloud_name and api_key and api_secret:
            cloudinary.config(
                cloud_name=cloud_name,
                api_key=api_key,
                api_secret=api_secret,
                secure=True
            )
            cloudinary_configured = True
except Exception as e:
    # En serverless, la configuración puede fallar silenciosamente si las vars no están disponibles
    # Se manejará más tarde cuando se intente usar Cloudinary
    cloudinary_configured = False

app = Flask(__name__)
app.config['SECRET_KEY'] = 'internlink_secret_2024'
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

JWT_SECRET = 'internlink2024'

# Ruta final de felicitación (hash fijo para no ser adivinable)
CONGRATS_PATH = hashlib.sha256(b"internlink_congratulations_2026").hexdigest()
YOUTUBE_REGALOS = [
    'https://www.youtube.com/watch?v=dQw4w9WgXcQ',
    'https://www.youtube.com/watch?v=EH3nK85MAIU',
    'https://www.youtube.com/watch?v=JlqW11vHLtw',
    'https://youtu.be/PNkhLZlsYjQ?si=jBnoxFNAcjBtiLpx',
    'https://youtu.be/jNhgozo9QJY?si=IZOMfUU5ik1vwo33',
    'https://www.youtube.com/watch?v=N5lTRsuUT5o',
    'https://www.youtube.com/watch?v=h69VanYG0Ds',
    'https://youtu.be/FQAcHm7-SoE?si=sJpPWuhufgPslFPG',
    'https://www.youtube.com/watch?v=AXp7ydbqTrw',
    'https://www.youtube.com/watch?v=tjiN9IYFutU',
    'https://www.youtube.com/watch?v=RUorAzaDftg',
]

# Cliente Redis Upstash con interfaz compatible con el resto de la app
_upstash = Redis.from_env()


class _Db:
    """Wrapper que adapta upstash_redis a la interfaz usada en la app."""

    def get(self, key):
        return _upstash.get(key)

    def set(self, key, value):
        return _upstash.set(key, value)

    def hgetall(self, key):
        out = _upstash.hgetall(key)
        return out if isinstance(out, dict) and out else {}

    def hset(self, key, field, value):
        return _upstash.hset(key, field, value)

    def hget(self, key, field):
        return _upstash.hget(key, field)

    def hmset(self, key, mapping):
        return _upstash.hset(key, values=mapping)

    def keys(self, pattern):
        out = _upstash.keys(pattern)
        return out if out is not None else []

    def delete(self, key):
        return _upstash.delete(key)

    def incr(self, key):
        return _upstash.incr(key)


db = _Db()

# Crear directorio de uploads solo si existe el sistema de archivos (desarrollo local)
# En Vercel/serverless, los archivos van a Cloudinary
try:
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
except (OSError, PermissionError):
    pass  # En serverless no hay sistema de archivos, se usa Cloudinary

try:
    os.makedirs('logs', exist_ok=True)
except (OSError, PermissionError):
    pass  # En serverless no se pueden crear directorios


# ---------------------------------------------------------------------------
# Datos iniciales
# ---------------------------------------------------------------------------
def init_db():
    """Poblar la base de datos con datos de ejemplo si están vacíos."""

    # Empresas
    if not db.hgetall('company:1'):
        db.hmset('company:1', {
            'id': '1',
            'name': 'TechCorp SA',
            'email': 'empresa@techcorp.com',
            'password': hashlib.sha256('EmpresaPass123!'.encode()).hexdigest(),
            'role': 'company',
            'sector': 'Tecnologia',
            'api_key': 'tc_prod_FLAG{session_hijacked}_v2',
            'contact_phone': '+57 604 555 0101',
        })
        db.hset('email_to_key', 'empresa@techcorp.com', 'company:1')

    if not db.hgetall('company:2'):
        db.hmset('company:2', {
            'id': '2',
            'name': 'DataFlow Inc',
            'email': 'rrhh@dataflow.com',
            'password': hashlib.sha256('DataFlow2024!'.encode()).hexdigest(),
            'role': 'company',
            'sector': 'Analisis de Datos',
            'api_key': 'df_prod_8a2c4e6f',
            'contact_phone': '+57 604 555 0202',
        })
        db.hset('email_to_key', 'rrhh@dataflow.com', 'company:2')

    if not db.hgetall('company:3'):
        db.hmset('company:3', {
            'id': '3',
            'name': 'SecureLog Corp',
            'email': 'rrhh@securelog.com',
            'password': hashlib.sha256('SecureLog99!'.encode()).hexdigest(),
            'role': 'company',
            'sector': 'Ciberseguridad',
            'api_key': 'sl_prod_a1b2c3d4',
            'contact_phone': '+57 604 555 0303',
        })
        db.hset('email_to_key', 'rrhh@securelog.com', 'company:3')

    # Sincronizar datos de empresas (por si ya existian sin api_key/sector/telefono)
    # Asegurar que company:1 tenga el rol correcto
    db.hset('company:1', 'role', 'company')
    db.hset('company:1', 'api_key', 'tc_prod_FLAG{session_hijacked}_v2')
    db.hset('company:1', 'sector', 'Tecnologia')
    db.hset('company:1', 'contact_phone', '+57 604 555 0101')
    db.hset('company:2', 'api_key', 'df_prod_8a2c4e6f')
    db.hset('company:2', 'sector', 'Analisis de Datos')
    db.hset('company:2', 'contact_phone', '+57 604 555 0202')
    db.hset('company:3', 'api_key', 'sl_prod_a1b2c3d4')
    db.hset('company:3', 'sector', 'Ciberseguridad')
    db.hset('company:3', 'contact_phone', '+57 604 555 0303')

    # Coordinador
    if not db.hgetall('coordinator:1'):
        db.hmset('coordinator:1', {
            'id': '1',
            'name': 'Carlos Ramirez',
            'email': 'coordinador@internlink.com',
            'password': hashlib.sha256('CoordPass123!'.encode()).hexdigest(),
            'role': 'coordinator',
            'department': 'Pasantias y Practicas',
        })
        db.hset('email_to_key', 'coordinador@internlink.com', 'coordinator:1')

    # Administrador
    if not db.hgetall('admin:1'):
        db.hmset('admin:1', {
            'id': '1',
            'name': 'Admin Sistema',
            'email': 'admin@internlink.com',
            'password': hashlib.sha256('AdminPass123!'.encode()).hexdigest(),
            'role': 'admin',
            'system_key': 'FLAG{internlink_compromised}',
            'admin_notes': 'Cuenta principal del sistema. Ref: FLAG{user_enumeration_is_real}',
        })
        db.hset('email_to_key', 'admin@internlink.com', 'admin:1')

    # Estudiantes de ejemplo
    if not db.hgetall('student:1'):
        db.hmset('student:1', {
            'id': '1',
            'name': 'Maria Gonzalez',
            'email': 'maria.gonzalez@mail.com',
            'password': hashlib.sha256('Student123!'.encode()).hexdigest(),
            'role': 'student',
            'university': 'Universidad de Medellin',
            'phone': '+57 300 111 2222',
            'bank_account': '13001523401',
        })
        db.hset('email_to_key', 'maria.gonzalez@mail.com', 'student:1')

    if not db.hgetall('student:2'):
        db.hmset('student:2', {
            'id': '2',
            'name': 'Andres Lopez',
            'email': 'andres.lopez@mail.com',
            'password': hashlib.sha256('Student456!'.encode()).hexdigest(),
            'role': 'student',
            'university': 'Universidad Nacional',
            'phone': '+57 300 333 4444',
            'bank_account': '13001523402',
        })
        db.hset('email_to_key', 'andres.lopez@mail.com', 'student:2')

    for sid, name, email, university, phone, bank in [
        ('3', 'Laura Martinez', 'laura.martinez@mail.com', 'Universidad de Antioquia', '+57 300 555 6666', '13001523403'),
        ('4', 'Carlos Rodriguez', 'carlos.rodriguez@mail.com', 'Universidad EAFIT', '+57 300 777 8888', '13001523404'),
        ('5', 'Patricia Mora', 'patricia.mora@mail.com', 'Universidad Pontificia', '+57 300 999 0000', '13001523405'),
        ('6', 'Felipe Restrepo', 'felipe.restrepo@mail.com', 'Universidad Nacional', '+57 300 111 3333', '13001523406'),
    ]:
        key = f'student:{sid}'
        if not db.hgetall(key):
            db.hmset(key, {
                'id': sid,
                'name': name,
                'email': email,
                'password': hashlib.sha256(('Student' + sid + '!').encode()).hexdigest(),
                'role': 'student',
                'university': university,
                'phone': phone,
                'bank_account': bank,
            })
            db.hset('email_to_key', email, key)

    # Sincronizar cuentas Bancolombia en todos los estudiantes (donde reciben el pago de pasantias)
    for sid in ['1', '2', '3', '4', '5', '6']:
        defaults = {'1': '13001523401', '2': '13001523402', '3': '13001523403', '4': '13001523404', '5': '13001523405', '6': '13001523406'}
        db.hset(f'student:{sid}', 'bank_account', defaults[sid])

    # Sincronizar indice email_to_key con usuarios iniciales (por si ya existian sin indice)
    for email, key in [
        ('empresa@techcorp.com', 'company:1'),
        ('rrhh@dataflow.com', 'company:2'),
        ('rrhh@securelog.com', 'company:3'),
        ('coordinador@internlink.com', 'coordinator:1'),
        ('admin@internlink.com', 'admin:1'),
        ('maria.gonzalez@mail.com', 'student:1'),
        ('andres.lopez@mail.com', 'student:2'),
        ('laura.martinez@mail.com', 'student:3'),
        ('carlos.rodriguez@mail.com', 'student:4'),
        ('patricia.mora@mail.com', 'student:5'),
        ('felipe.restrepo@mail.com', 'student:6'),
    ]:
        db.hset('email_to_key', email, key)

    # Asignaciones de pasantía (intern records)
    if not db.hgetall('intern:1'):
        db.hmset('intern:1', {
            'id': '1',
            'student_id': '1',
            'student_name': 'Maria Gonzalez',
            'student_email': 'maria.gonzalez@mail.com',
            'company_id': '1',
            'company_name': 'TechCorp SA',
            'position': 'Desarrolladora Backend',
            'salary': '1500000',
            'status': 'active',
            'start_date': '2026-01-15',
            'evaluation': 'Buen rendimiento general.',
            'cv_path': '',
        })

    if not db.hgetall('intern:2'):
        db.hmset('intern:2', {
            'id': '2',
            'student_id': '2',
            'student_name': 'Andres Lopez',
            'student_email': 'andres.lopez@mail.com',
            'company_id': '2',
            'company_name': 'DataFlow Inc',
            'position': 'Analista de Datos',
            'salary': '1800000',
            'status': 'active',
            'start_date': '2026-02-01',
            'evaluation': 'Calificacion 9.5/10. Buen desempeno en analitica.',
            'cv_path': '',
        })

    if not db.hgetall('intern:3'):
        db.hmset('intern:3', {
            'id': '3',
            'student_id': '3',
            'student_name': 'Laura Martinez',
            'student_email': 'laura.martinez@mail.com',
            'company_id': '1',
            'company_name': 'TechCorp SA',
            'position': 'Disenadora UX/UI',
            'salary': '1600000',
            'status': 'pending',
            'start_date': '2026-02-10',
            'evaluation': 'Pendiente de evaluacion.',
            'cv_path': '',
        })

    if not db.hgetall('intern:4'):
        db.hmset('intern:4', {
            'id': '4',
            'student_id': '4',
            'student_name': 'Carlos Rodriguez',
            'student_email': 'carlos.rodriguez@mail.com',
            'company_id': '2',
            'company_name': 'DataFlow Inc',
            'position': 'Ingeniero de Datos Junior',
            'salary': '1700000',
            'status': 'active',
            'start_date': '2026-01-20',
            'evaluation': 'Buen desempeno en el primer mes.',
            'cv_path': '',
        })

    if not db.hgetall('intern:5'):
        db.hmset('intern:5', {
            'id': '5',
            'student_id': '5',
            'student_name': 'Patricia Mora',
            'student_email': 'patricia.mora@mail.com',
            'company_id': '3',
            'company_name': 'SecureLog Corp',
            'position': 'Analista de Seguridad',
            'salary': '2200000',
            'status': 'active',
            'start_date': '2026-01-10',
            'evaluation': 'Evaluacion confidencial. Codigo de auditoria: FLAG{idor_horizontal}',
            'cv_path': '',
        })

    if not db.hgetall('intern:6'):
        db.hmset('intern:6', {
            'id': '6',
            'student_id': '6',
            'student_name': 'Felipe Restrepo',
            'student_email': 'felipe.restrepo@mail.com',
            'company_id': '3',
            'company_name': 'SecureLog Corp',
            'position': 'Pentester Junior',
            'salary': '2000000',
            'status': 'active',
            'start_date': '2026-01-05',
            'evaluation': 'Rendimiento adecuado en pruebas de intrusion. Ref: validar desde perfil estudiante; endpoint perfil acepta campos extendidos.',
            'cv_path': '',
        })

    # Sincronizar interns 1-4 (por si ya existian con tildes o con flag en company 2)
    db.hset('intern:1', 'student_name', 'Maria Gonzalez')
    db.hset('intern:2', 'student_name', 'Andres Lopez')
    db.hset('intern:2', 'evaluation', 'Calificacion 9.5/10. Buen desempeno en analitica.')
    db.hset('intern:3', 'student_name', 'Laura Martinez')
    db.hset('intern:3', 'evaluation', 'Pendiente de evaluacion.')
    db.hset('intern:4', 'student_name', 'Carlos Rodriguez')
    db.hset('intern:4', 'evaluation', 'Buen desempeno en el primer mes.')
    db.hset('intern:6', 'evaluation', 'Rendimiento adecuado en pruebas de intrusion. Ref: validar desde perfil estudiante; endpoint perfil acepta campos extendidos.')

    # Ofertas de pasantía
    if not db.hgetall('offer:1'):
        db.hmset('offer:1', {
            'id': '1',
            'company_id': '1',
            'company_name': 'TechCorp SA',
            'title': 'Desarrollador Backend Junior',
            'description': 'Desarrollo de APIs REST con Python y Flask. Se requiere conocimiento de bases de datos.',
            'salary': '1500000',
            'status': 'active',
            'created_at': '2026-01-20',
        })

    if not db.hgetall('offer:2'):
        db.hmset('offer:2', {
            'id': '2',
            'company_id': '2',
            'company_name': 'DataFlow Inc',
            'title': 'Analista de Datos',
            'description': 'Analisis y visualizacion de datos empresariales. Excel y Python requeridos.',
            'salary': '1800000',
            'status': 'active',
            'created_at': '2026-01-25',
        })

    if not db.hgetall('offer:3'):
        db.hmset('offer:3', {
            'id': '3',
            'company_id': '1',
            'company_name': 'TechCorp SA',
            'title': 'Disenador UX/UI',
            'description': 'Diseno de interfaces y experiencia de usuario para aplicaciones web.',
            'salary': '1600000',
            'status': 'pending',
            'created_at': '2026-02-01',
        })

    # Avisos del sistema (visibles para coordinadores)
    if not db.hgetall('notice:1'):
        db.hmset('notice:1', {
            'id': '1',
            'title': 'Actualizacion del modulo de exportacion',
            'message': 'Los reportes de candidatos ahora incluyen informacion de evaluacion y salarios.',
            'date': '2026-02-05',
            'type': 'info',
        })

    if not db.hgetall('notice:2'):
        db.hmset('notice:2', {
            'id': '2',
            'title': 'Mantenimiento programado',
            'message': 'El sistema estara en mantenimiento el sabado 22 de febrero de 2:00 AM a 6:00 AM.',
            'date': '2026-02-10',
            'type': 'warning',
        })

    if not db.hgetall('notice:3'):
        db.hmset('notice:3', {
            'id': '3',
            'title': 'Auditoria de seguridad',
            'message': 'Se detectaron cambios no autorizados en perfiles de usuario. Token de auditoria: FLAG{mass_assignment_abuse}',
            'date': '2026-01-28',
            'type': 'security',
        })

    # Configuración del sistema (sin master_key: esa flag está solo en el archivo de logs)
    if not db.hgetall('system:config'):
        db.hmset('system:config', {
            'version': '2.1.0',
            'jwt_secret': 'internlink2024',
            'admin_verification': 'FLAG{jwt_forged_successfully}',
            'maintenance_mode': 'false',
            'max_upload_size': '16MB',
            'debug_log_path': '/logs/debug.log',
        })
    db.hset('system:config', 'debug_log_path', '/logs/debug.log')


# ---------------------------------------------------------------------------
# Autenticación
# ---------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        session_token = request.cookies.get('session_token')
        if not session_token:
            return redirect(url_for('login'))
        user_data = db.get(f'session:{session_token}')
        if not user_data:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def get_current_user():
    session_token = request.cookies.get('session_token')
    if not session_token:
        return None
    user_json = db.get(f'session:{session_token}')
    if not user_json:
        return None
    try:
        return json.loads(user_json)
    except:
        return None


def redirect_by_role(user):
    """Redirige al usuario a su dashboard correspondiente según su rol.
    Importante para seguridad: si se roba una cookie, el usuario será redirigido
    automáticamente a su dashboard correcto según su rol real."""
    if not user:
        return redirect(url_for('index'))
    role = user.get('role', 'student')
    routes = {
        'student': 'student_dashboard',
        'company': 'company_dashboard',
        'coordinator': 'coordinator_dashboard',
        'admin': 'admin_dashboard',
    }
    dashboard_route = routes.get(role, 'student_dashboard')
    return redirect(url_for(dashboard_route))


def get_admin_user():
    """Verificar acceso de administrador vía JWT o sesión."""
    # JWT vía cookie admin_token
    token = request.cookies.get('admin_token')
    if not token:
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            token = auth_header[7:]

    if token:
        try:
            decoded = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
            if decoded.get('role') == 'admin':
                return decoded
        except jwt.InvalidTokenError:
            pass
        # Fallback: intentar sin verificación de firma
        try:
            decoded = jwt.decode(token, options={"verify_signature": False})
            if decoded.get('role') == 'admin':
                return decoded
        except Exception:
            pass

    # Sesión convencional
    user = get_current_user()
    if user and user.get('role') == 'admin':
        return user
    return None


def _normalize_email(email):
    """Misma normalización en registro y login para que el índice coincida."""
    if not email or not isinstance(email, str):
        return ''
    return email.strip().lower()


def check_email_exists(email):
    """Verifica si un email existe usando el índice email_to_key (optimizado)."""
    email = _normalize_email(email)
    if not email:
        return False
    key = db.hget('email_to_key', email)
    return key is not None and key != ''


def find_user_by_email(email):
    """Buscar un usuario por email usando el índice email_to_key (optimizado)."""
    email = _normalize_email(email)
    if not email:
        return None, None
    key = db.hget('email_to_key', email)
    if not key:
        # Fallback: índice puede tener el email con otra capitalización (usuarios ya registrados)
        index = db.hgetall('email_to_key') or {}
        for stored_email, stored_key in index.items():
            if _normalize_email(stored_email) == email:
                key = stored_key
                break
    if not key:
        return None, None
    key = str(key).strip()
    data = db.hgetall(key)
    if not data:
        return None, None
    data['user_key'] = key
    return data, key


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
def log_entry(message):
    """Escribe entrada de log. En serverless (Vercel), los archivos no persisten entre invocaciones,
    pero esto está bien porque create_debug_log() regenera el archivo en cada inicio."""
    log_path = 'logs/debug.log'
    ts = datetime.now().strftime('[%Y-%m-%d %H:%M:%S]')
    try:
        # Crear directorio si no existe (puede fallar en serverless, pero lo intentamos)
        os.makedirs('logs', exist_ok=True)
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(f"{ts} INFO: {message}\n")
    except (OSError, IOError) as e:
        # En serverless, escribir archivos puede fallar silenciosamente
        # Los logs se perderán, pero la aplicación seguirá funcionando
        pass


def create_debug_log():
    """Crea/regenera el archivo de log de debug. En serverless (Vercel), esto guarda el contenido
    en Redis en lugar del sistema de archivos."""
    content = (
        "[2026-01-10 08:15:32] INFO: Servidor iniciado en puerto 5000\n"
        "[2026-01-10 08:15:33] DEBUG: Conexión a Redis establecida\n"
        "[2026-01-10 08:16:01] INFO: Login de usuario: admin@internlink.com (role: admin)\n"
        "[2026-01-10 08:17:45] WARNING: Intento de login fallido para admin@internlink.com\n"
        "[2026-01-10 08:18:02] DEBUG: Cargando configuración JWT...\n"
        "[2026-01-10 08:18:02] DEBUG: JWT_SECRET = internlink2024\n"
        "[2026-01-10 08:18:03] INFO: Módulo JWT inicializado\n"
        "[2026-01-10 08:20:15] ERROR: Excepción no controlada en /api/company/candidates\n"
        "    Traceback (most recent call last):\n"
        "      File \"app.py\", line 342, in get_candidates\n"
        "        validate_company_ownership(user_id, company_id)\n"
        "    AttributeError: module has no attribute 'validate_company_ownership'\n"
        "[2026-01-10 08:22:30] DEBUG: Credenciales temporales creadas: temp_admin / TempPass2026!\n"
        "[2026-01-10 08:25:00] INFO: Token de auditoría del sistema: FLAG{logs_are_sensitive}\n"
        "[2026-01-10 08:30:00] DEBUG: Admin auth configurado: JWT via cookie 'admin_token'\n"
        "[2026-01-10 08:35:00] INFO: Proceso de backup completado\n"
        "[2026-01-10 08:36:00] INFO: Actualizacion masiva de cuentas Bancolombia (pago de pasantias) — endpoint /api/admin/bulk-update-payment-accounts — solo admin\n"
        "[2026-01-10 08:40:00] WARNING: Rate limiting no configurado para /api/check-email\n"
        "[2026-01-10 08:45:00] INFO: Bot de revision de CV iniciado — sesion: rev_bot_2026\n"
    )
    
    # Guardar en Redis para acceso en Vercel/serverless
    try:
        db.set('system:debug_log', content)
    except Exception:
        pass
    
    # Intentar guardar también en archivo (solo funciona en local)
    try:
        os.makedirs('logs', exist_ok=True)
        log_path = 'logs/debug.log'
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write(content)
    except (OSError, IOError):
        # En serverless, crear archivos puede fallar. Se usa Redis en su lugar.
        pass


# ---------------------------------------------------------------------------
# Rutas públicas
# ---------------------------------------------------------------------------
@app.route('/health')
def health():
    """Health check para debugging en Vercel."""
    import platform
    try:
        # Verificar conexión a Redis
        redis_ok = False
        try:
            db.set('health_check', 'ok')
            redis_ok = db.get('health_check') == 'ok'
        except Exception as e:
            redis_ok = f"Error: {str(e)}"
        
        # Verificar Cloudinary
        cloudinary_ok = bool(os.getenv('CLOUDINARY_URL'))
        
        info = {
            'status': 'ok',
            'python_version': platform.python_version(),
            'flask_version': '3.0.0',
            'redis': redis_ok,
            'cloudinary': cloudinary_ok,
            'templates_folder': app.template_folder,
            'static_folder': app.static_folder,
            'cwd': os.getcwd(),
            'templates_exist': os.path.exists('templates'),
            'static_exist': os.path.exists('static'),
        }
        return jsonify(info), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/')
def index():
    user = get_current_user()
    if user:
        role = user.get('role', 'student')
        routes = {
            'company': 'company_dashboard',
            'coordinator': 'coordinator_dashboard',
            'admin': 'admin_dashboard',
        }
        return redirect(url_for(routes.get(role, 'student_dashboard')))
    return render_template('gate.html')


@app.route('/<hash_segment>/congratulations')
def congratulations(hash_segment):
    """Ruta final de felicitación tras completar el lab (hash fijo en la URL)."""
    if hash_segment != CONGRATS_PATH:
        abort(404)
    return render_template('congratulations.html', youtube_regalos=YOUTUBE_REGALOS)


def email_verification_required(f):
    """Exige haber verificado un correo (paso previo) para acceder a login o registro."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if request.cookies.get('email_verified') != '1':
            return redirect(url_for('index', msg='verify_first'))
        return f(*args, **kwargs)
    return decorated


@app.route('/login', methods=['GET', 'POST'])
@email_verification_required
def login():
    if request.method == 'POST':
        email = _normalize_email(request.form.get('email'))
        password = (request.form.get('password') or '').strip()

        user, user_key = find_user_by_email(email)

        if user and password:
            stored_password = str(user.get('password') or '')
            expected_hash = hashlib.sha256(password.encode()).hexdigest()
            if stored_password == expected_hash:
                session_token = secrets.token_hex(32)
                user['user_key'] = user_key
                db.set(f'session:{session_token}', json.dumps(user))

                log_entry(f"Login exitoso: {email} — Role: {user.get('role')}")

                resp = make_response(redirect(url_for('index')))
                resp.set_cookie('session_token', session_token, httponly=False)
                return resp

        return render_template('login.html', error='Correo o contraseña incorrectos', email=request.form.get('email'))

    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
@email_verification_required
def register():
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        email = _normalize_email(request.form.get('email'))
        password = (request.form.get('password') or '').strip()
        role = (request.form.get('role') or 'student').strip().lower()

        if not email or not password:
            return render_template('register.html', error='Correo y contraseña son obligatorios', email=email or request.form.get('email'))

        if check_email_exists(email):
            return render_template('register.html', error='Este correo electrónico ya está registrado')

        user_id = str(db.incr(f'counter:{role}') or 1)
        password_hash = hashlib.sha256(password.encode()).hexdigest()

        user_data = {
            'id': user_id,
            'name': name,
            'email': email,
            'password': password_hash,
            'role': role,
            'created_at': datetime.now().isoformat(),
        }
        user_key = f'{role}:{user_id}'
        db.hmset(user_key, user_data)
        db.hset('email_to_key', email, user_key)

        log_entry(f"Nuevo registro: {email} — Role: {role}")
        return redirect(url_for('login'))

    return render_template('register.html', email=request.args.get('email', ''))


# Solo las cuentas guardadas como coordinator:* o admin:* en Redis pueden invitar (por clave, no por campo role)
def _key_can_invite(redis_key):
    """Determina por la clave Redis si el usuario puede invitar (solo coordinator y admin por tipo de cuenta)."""
    return redis_key.startswith('coordinator:') or redis_key.startswith('admin:')


def get_registered_users_with_invite_flag():
    """Lista de usuarios registrados. Mal implementado: devuelve todos con un campo que filtra quién puede invitar.
    Optimizado: usa el índice email_to_key."""
    email_to_key = db.hgetall('email_to_key') or {}
    users = []
    for email, key in email_to_key.items():
        puede_invitar = _key_can_invite(key)
        users.append({'email': email, 'puede_invitar': puede_invitar})
    return users


@app.route('/api/check-email', methods=['GET', 'POST'])
def check_email():
    if request.method == 'GET':
        # GET mal implementado: estaba pensado para devolver solo invitadores pero filtra mal y devuelve todos los registrados
        users = get_registered_users_with_invite_flag()
        return jsonify({'usuarios': users}), 200

    # POST: verificar si el correo corresponde a un invitador válido
    email = request.json.get('email') if request.is_json else None
    if not email:
        return jsonify({'error': 'Email requerido'}), 400

    user, user_key = find_user_by_email(email)

    if not user:
        return jsonify({
            'valid_inviter': False,
            'message': 'Este correo no está registrado en el sistema.',
        }), 200

    if not _key_can_invite(user_key):
        return jsonify({
            'valid_inviter': False,
            'message': 'Este usuario no tiene permiso para invitar a otros. Solo algunos roles pueden hacerlo.',
        }), 200

    # Es un invitador válido: desbloquear acceso y devolver flag en header
    resp = jsonify({
        'valid_inviter': True,
        'message': 'Invitación verificada correctamente. Ya puede acceder al sistema.',
    })
    resp.headers['X-Request-Id'] = 'FLAG{user_enumeration_is_real}'
    resp.set_cookie('email_verified', '1', max_age=300, path='/')
    return resp, 200


@app.route('/logout')
def logout():
    session_token = request.cookies.get('session_token')
    if session_token:
        db.delete(f'session:{session_token}')
    resp = make_response(redirect(url_for('index')))
    resp.set_cookie('session_token', '', expires=0)
    resp.set_cookie('admin_token', '', expires=0)
    resp.set_cookie('email_verified', '', expires=0)
    return resp


# ---------------------------------------------------------------------------
# Panel estudiante
# ---------------------------------------------------------------------------
@app.route('/dashboard/student')
@login_required
def student_dashboard():
    user = get_current_user()
    # Validar que el usuario sea estudiante, si no, redirigir a su dashboard correspondiente
    if user.get('role') != 'student':
        return redirect_by_role(user)
    # Cargar ofertas activas
    offers = []
    for key in db.keys('offer:*'):
        offer = db.hgetall(key)
        if offer.get('status') == 'active':
            offers.append(offer)
    return render_template('student_dashboard.html', user=user, offers=offers)


@app.route('/upload-cv', methods=['POST'])
@login_required
def upload_cv():
    if 'cv' not in request.files:
        return jsonify({'error': 'No se recibió ningún archivo'}), 400

    file = request.files['cv']
    if file.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío'}), 400

    if not file.filename.endswith('.pdf'):
        return jsonify({'error': 'Solo se permiten archivos en formato PDF'}), 400

    filename = secure_filename(file.filename)
    user = get_current_user()
    user_filename = f"{user['id']}_{filename}"
    
    # Subir a Cloudinary (resource_type="raw" para preservar el contenido exacto, incluyendo HTML)
    cloudinary_url = None
    cloudinary_public_id = None
    has_cloudinary_url = bool(os.getenv('CLOUDINARY_URL'))
    has_cloudinary_keys = bool(os.getenv('CLOUDINARY_CLOUD_NAME') and os.getenv('CLOUDINARY_API_KEY'))
    cloudinary_configured = has_cloudinary_url or has_cloudinary_keys
    
    log_entry(f"Cloudinary check: URL={has_cloudinary_url}, Keys={has_cloudinary_keys}, Configurado={cloudinary_configured}")
    
    if cloudinary_configured:
        try:
            file_content = file.read()
            file.seek(0)
            
            log_entry(f"Intentando subir a Cloudinary: {user_filename} (tamaño: {len(file_content)} bytes)")
            # Subir archivo como público explícitamente
            # IMPORTANTE: No usar 'folder' junto con 'public_id' que ya incluye la ruta
            # Usar solo public_id con la ruta completa para evitar problemas
            public_id_full = f"internlink/cvs/{user_filename}"
            
            # Intentar eliminar el archivo anterior si existe (para evitar problemas con archivos privados antiguos)
            try:
                cloudinary.uploader.destroy(public_id_full, resource_type="raw", invalidate=True)
                log_entry(f"Archivo anterior eliminado (si existía): {public_id_full}")
            except Exception as del_err:
                # No es crítico si no existe
                log_entry(f"No se pudo eliminar archivo anterior (puede que no exista): {str(del_err)}")
            
            upload_result = cloudinary.uploader.upload(
                file_content,
                public_id=public_id_full,  # Incluir la ruta completa en public_id
                resource_type="raw",  # "raw" preserva el contenido sin transformaciones (importante para XSS)
                overwrite=True,
                type="upload",  # type="upload" hace el archivo público por defecto
                invalidate=True,  # Invalidar caché
            )
            cloudinary_url = upload_result['secure_url']
            # Guardar también el public_id completo para poder descargarlo después
            cloudinary_public_id = upload_result.get('public_id', public_id_full)
            
            # Generar URL firmada para bypassear restricciones de cuentas gratuitas (PDFs bloqueados por defecto)
            # Las URLs firmadas permiten acceso incluso si la cuenta tiene restricciones
            signed_url = cloudinary_utils.cloudinary_url(
                cloudinary_public_id,
                resource_type="raw",
                secure=True,
                type="upload",
                sign_url=True  # Firmar la URL para bypassear restricciones
            )[0]
            
            log_entry(f"CV subido exitosamente a Cloudinary: {user_filename} -> {cloudinary_url} (public_id: {cloudinary_public_id})")
            log_entry(f"URL pública: {cloudinary_url}")
            log_entry(f"URL firmada (bypass restricciones): {signed_url}")
            log_entry(f"Tipo de entrega en respuesta: {upload_result.get('type', 'N/A')}")
            
            # Usar la URL firmada para almacenar (bypassea restricciones de cuentas gratuitas)
            cloudinary_url = signed_url
        except Exception as e:
            error_msg = str(e)
            import traceback
            log_entry(f"ERROR al subir a Cloudinary: {error_msg}\n{traceback.format_exc()}")
            # Si Cloudinary está configurado pero falla, NO hacer fallback - mostrar error
            return jsonify({'error': f'Error al subir el archivo a Cloudinary: {error_msg}. Verifica tu configuración.'}), 500
    
    # Fallback solo si Cloudinary NO está configurado (desarrollo local sin Cloudinary)
    if not cloudinary_url and not cloudinary_configured:
        if os.path.exists(app.config['UPLOAD_FOLDER']):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], user_filename)
            file.save(filepath)
            cloudinary_url = f"/view-cv/{user_filename}"
            log_entry(f"CV guardado localmente (Cloudinary no configurado): {user_filename}")
        else:
            return jsonify({'error': 'Cloudinary no configurado y no hay sistema de archivos disponible'}), 500
    elif not cloudinary_url:
        # Cloudinary estaba configurado pero no se obtuvo URL (no debería llegar aquí)
        return jsonify({'error': 'Error: Cloudinary configurado pero no se obtuvo URL'}), 500

    # Guardar URL de Cloudinary en el perfil del usuario
    user_key = user.get('user_key', f"{user['role']}:{user['id']}")
    db.hset(user_key, 'cv_path', cloudinary_url)
    db.hset(user_key, 'cv_filename', user_filename)  # Guardar también el filename para referencia
    # Guardar public_id de Cloudinary para poder descargarlo con la API si es necesario
    if cloudinary_public_id:
        db.hset(user_key, 'cv_cloudinary_id', cloudinary_public_id)

    # Procesar CV de forma automática (pasa la URL de Cloudinary y el public_id si está disponible)
    process_cv(cloudinary_url, user_filename)

    return jsonify({
        'success': True,
        'message': 'CV subido correctamente. Será revisado próximamente.',
        'filename': user_filename,
    })


def _extract_webhook_from_html(file_url_or_path, cloudinary_public_id=None):
    """Extrae la primera URL tipo webhook (http(s)) del archivo, para simular exfiltración.
    file_url_or_path puede ser una URL de Cloudinary, una ruta local, o una ruta relativa /view-cv/...
    cloudinary_public_id: si es una URL de Cloudinary, usar la API para descargar en lugar de HTTP directo."""
    try:
        # Si es una URL de Cloudinary, descargar usando la URL (puede ser firmada o pública)
        if file_url_or_path.startswith('https://res.cloudinary.com/') or file_url_or_path.startswith('http://res.cloudinary.com/'):
            # Si la URL ya está firmada (contiene /s--), usarla directamente
            # Si no está firmada y tenemos public_id, generar URL firmada para bypassear restricciones
            try:
                log_entry(f"Descargando archivo desde Cloudinary: {file_url_or_path}")
                # Usar Request con headers para evitar problemas de autenticación
                req = Request(file_url_or_path)
                req.add_header('User-Agent', 'Mozilla/5.0')
                response = urlopen(req, timeout=10)
                content_bytes = response.read()
                log_entry(f"Archivo descargado exitosamente desde Cloudinary: {len(content_bytes)} bytes")
            except Exception as e:
                log_entry(f"Error al descargar desde Cloudinary: {str(e)}")
                # Si falla y tenemos public_id, generar URL firmada para bypassear restricciones
                if cloudinary_public_id:
                    try:
                        # Generar URL firmada para bypassear restricciones de cuentas gratuitas
                        signed_url = cloudinary_utils.cloudinary_url(
                            cloudinary_public_id, 
                            resource_type="raw", 
                            secure=True,
                            type="upload",
                            sign_url=True  # Firmar la URL para bypassear restricciones
                        )[0]
                        log_entry(f"Intentando descargar con URL firmada (bypass restricciones): {signed_url}")
                        req2 = Request(signed_url)
                        req2.add_header('User-Agent', 'Mozilla/5.0')
                        response = urlopen(req2, timeout=10)
                        content_bytes = response.read()
                        log_entry(f"Archivo descargado usando URL firmada: {len(content_bytes)} bytes")
                    except Exception as url_err:
                        log_entry(f"Error también con URL firmada: {str(url_err)}")
                        return None
                else:
                    return None
        elif file_url_or_path.startswith('http://') or file_url_or_path.startswith('https://'):
            # Otra URL HTTP (no Cloudinary)
            response = urlopen(file_url_or_path, timeout=10)
            content_bytes = response.read()
        elif file_url_or_path.startswith('/view-cv/'):
            # Ruta relativa /view-cv/filename -> convertir a ruta local
            filename = file_url_or_path.replace('/view-cv/', '')
            local_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            if os.path.exists(local_path):
                with open(local_path, 'rb') as f:
                    content_bytes = f.read()
            else:
                return None
        else:
            # Ruta local absoluta (fallback para desarrollo)
            if os.path.exists(file_url_or_path):
                with open(file_url_or_path, 'rb') as f:
                    content_bytes = f.read()
            else:
                return None
        
        # Verificar si es PDF real
        if content_bytes[:5] == b'%PDF-':
            return None
        
        # Leer como texto
        content = content_bytes.decode('utf-8', errors='ignore')
        
        # Buscar URLs http(s) en el contenido (p. ej. var w = "https://webhook.site/...")
        match = re.search(r'https?://[^\s"\'<>)\];]+', content)
        if not match:
            return None
        url = match.group(0).rstrip('.,;:')
        # No enviar a localhost ni a nosotros mismos
        if 'localhost' in url or '127.0.0.1' in url:
            return None
        return url
    except Exception as e:
        log_entry(f"Error al extraer webhook de {file_url_or_path}: {str(e)}")
        return None


def process_cv(file_url, filename):
    """Proceso automático que simula la revisión del CV por parte del sistema (empresa abre el CV).
    file_url: URL de Cloudinary o ruta local."""
    company = db.hgetall('company:1')
    if not company:
        log_entry(f"ERROR: company:1 no encontrada en Redis")
        return
    
    # Verificar que company:1 tenga el rol correcto antes de crear la sesión
    if company.get('role') != 'company':
        log_entry(f"ADVERTENCIA: company:1 tiene rol incorrecto: {company.get('role')}, corrigiendo a 'company'")
        db.hset('company:1', 'role', 'company')
        company = db.hgetall('company:1')  # Re-obtener después de corregir
    
    bot_session = secrets.token_hex(32)
    # Crear sesión con datos explícitos de empresa, asegurando el rol correcto
    company_data = {
        'id': company.get('id', '1'),
        'name': company.get('name', 'TechCorp SA'),
        'email': company.get('email', 'empresa@techcorp.com'),
        'role': 'company',  # FORZAR rol de empresa explícitamente
        'sector': company.get('sector', 'Tecnologia'),
        'api_key': company.get('api_key', ''),
        'contact_phone': company.get('contact_phone', ''),
        'user_key': 'company:1',
    }
    # Guardar sesión con datos explícitos de empresa
    db.set(f'session:{bot_session}', json.dumps(company_data))
    log_entry(f"Sesión de bot creada para empresa: role={company_data.get('role')}, id={company_data.get('id')}, email={company_data.get('email')}, token={bot_session[:16]}...")

    review_id = str(db.incr('counter:review') or 1)
    db.hmset(f'review:{review_id}', {
        'id': review_id,
        'filename': filename,
        'file_url': file_url,
        'status': 'processed',
        'reviewer': 'system_bot',
        'session_token': bot_session,
        'review_note': 'Archivo procesado correctamente. Ref: FLAG{stored_xss_persisted}',
        'processed_at': datetime.now().isoformat(),
    })

    log_entry(f"CV revisado automáticamente — Sesión del bot: {bot_session} — Archivo: {filename} — URL: {file_url}")

    # Buscar el public_id de Cloudinary en Redis (si existe)
    cloudinary_public_id = None
    for key in db.keys('student:*'):
        user = db.hgetall(key)
        if user.get('cv_filename') == filename:
            cloudinary_public_id = user.get('cv_cloudinary_id')
            break

    # Simular que la empresa abrió el CV: si el archivo es HTML con webhook, enviar la cookie de la empresa
    webhook = _extract_webhook_from_html(file_url, cloudinary_public_id)
    log_entry(f"Webhook extraído del CV: {webhook if webhook else 'No encontrado'}")
    if webhook:
        flag_act2 = 'FLAG{stored_xss_persisted}'
        cookie_value = f"session_token={bot_session}"
        exfil_url = f"{webhook.rstrip('/')}?c={quote(cookie_value)}&flag={quote(flag_act2)}"
        try:
            req = Request(exfil_url, headers={'User-Agent': 'Mozilla/5.0 (compatible; InternLink/1.0)'})
            urlopen(req, timeout=5)
        except (URLError, OSError):
            pass


@app.route('/view-cv/<filename>')
def view_cv(filename):
    """Servir el CV: PDF real se muestra con visor nativo; otro contenido (p. ej. HTML) como HTML.
    Si el archivo está en Cloudinary, redirige o sirve desde ahí."""
    # Buscar el CV en Redis por filename
    cv_url = None
    for key in db.keys('student:*'):
        user = db.hgetall(key)
        if user.get('cv_filename') == filename:
            cv_url = user.get('cv_path')
            break
    
    # Si encontramos URL de Cloudinary, redirigir o servir desde ahí
    if cv_url and (cv_url.startswith('http://') or cv_url.startswith('https://')):
        # Para HTML malicioso: servir directamente desde Cloudinary para que el JS se ejecute
        # Cloudinary sirve archivos raw sin transformaciones si usamos resource_type="raw"
        try:
            response = urlopen(cv_url, timeout=10)
            content = response.read()
            
            # Si es PDF real, servir como PDF
            if content[:5] == b'%PDF-':
                return send_file(
                    BytesIO(content),
                    mimetype='application/pdf',
                    as_attachment=False,
                    download_name=filename,
                )
            # Si es HTML, servir como HTML (importante para XSS)
            return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
        except Exception as e:
            # Si falla, intentar con URL firmada (bypass restricciones de cuentas gratuitas)
            log_entry(f"Error al cargar desde URL pública: {str(e)}, intentando con URL firmada")
            try:
                # Buscar public_id en Redis
                cloudinary_public_id = None
                for key in db.keys('student:*'):
                    user = db.hgetall(key)
                    if user.get('cv_filename') == filename:
                        cloudinary_public_id = user.get('cv_cloudinary_id')
                        break
                
                if cloudinary_public_id:
                    # Generar URL firmada para bypassear restricciones
                    signed_url = cloudinary_utils.cloudinary_url(
                        cloudinary_public_id,
                        resource_type="raw",
                        secure=True,
                        type="upload",
                        sign_url=True
                    )[0]
                    log_entry(f"Usando URL firmada: {signed_url}")
                    response = urlopen(signed_url, timeout=10)
                    content = response.read()
                    
                    # Si es PDF real, servir como PDF
                    if content[:5] == b'%PDF-':
                        return send_file(
                            BytesIO(content),
                            mimetype='application/pdf',
                            as_attachment=False,
                            download_name=filename,
                        )
                    # Si es HTML, servir como HTML (importante para XSS)
                    return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
                else:
                    return "Error al cargar el archivo: no se encontró public_id", 500
            except Exception as e2:
                log_entry(f"Error también con URL firmada: {str(e2)}")
                return f"Error al cargar el archivo: {str(e2)}", 500
    
    # Fallback: buscar localmente (solo para desarrollo)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        with open(filepath, 'rb') as f:
            header = f.read(5)
        if header == b'%PDF-':
            return send_file(
                filepath,
                mimetype='application/pdf',
                as_attachment=False,
                download_name=filename,
            )
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    
    return "Archivo no encontrado", 404


@app.route('/api/profile/update', methods=['PUT'])
@login_required
def update_profile():
    user = get_current_user()
    data = request.json
    if not data:
        return jsonify({'error': 'Datos requeridos'}), 400

    # No permitir ascender a admin por mass assignment (el lab exige coordinator -> JWT para admin)
    if data.get('role') == 'admin' and user.get('role') != 'admin':
        return jsonify({'error': 'El rol admin no se puede asignar desde este endpoint'}), 403

    user_key = user.get('user_key', f"{user['role']}:{user['id']}")

    for field, value in data.items():
        db.hset(user_key, field, str(value))

    # Actualizar la sesión con los datos nuevos
    session_token = request.cookies.get('session_token')
    updated = db.hgetall(user_key)
    updated['user_key'] = user_key
    db.set(f'session:{session_token}', json.dumps(updated))

    return jsonify({'success': True, 'message': 'Perfil actualizado correctamente'})


# ---------------------------------------------------------------------------
# Ofertas públicas
# ---------------------------------------------------------------------------
@app.route('/api/offers')
@login_required
def get_offers():
    offers = []
    for key in db.keys('offer:*'):
        offer = db.hgetall(key)
        if offer.get('status') == 'active':
            offers.append(offer)
    return jsonify({'offers': offers})


# ---------------------------------------------------------------------------
# Panel empresa
# ---------------------------------------------------------------------------
@app.route('/dashboard/company')
@login_required
def company_dashboard():
    user = get_current_user()
    # Validar que el usuario sea empresa, si no, redirigir a su dashboard correspondiente
    if user.get('role') != 'company':
        return redirect_by_role(user)

    # Actualizar usuario con datos frescos de Redis (api_key, sector, teléfono) por si la sesión es antigua
    user_key = user.get('user_key') or f"company:{user.get('id')}"
    fresh = db.hgetall(user_key)
    if fresh:
        user = dict(user)
        user.update({k: v for k, v in fresh.items() if k not in ('password',)})

    # Ofertas de esta empresa
    offers = []
    for key in db.keys('offer:*'):
        offer = db.hgetall(key)
        if offer.get('company_id') == user.get('id'):
            offers.append(offer)

    return render_template('company_dashboard.html', user=user, offers=offers)


@app.route('/internal/docs/securelog-corp')
def internal_docs_securelog():
    """Documentacion interna SecureLog Corp (referenciada en respuestas confidenciales)."""
    base_url = request.url_root
    return render_template('internal_docs.html', base_url=base_url)


@app.route('/api/company/candidates')
@login_required
def get_candidates():
    company_id = request.args.get('company_id')
    if not company_id:
        user = get_current_user()
        company_id = user.get('id')

    candidates = []
    for key in db.keys('intern:*'):
        intern = db.hgetall(key)
        if intern.get('company_id') == company_id:
            candidates.append({
                'id': intern.get('id'),
                'student_name': intern.get('student_name'),
                'student_email': intern.get('student_email'),
                'position': intern.get('position'),
                'salary': intern.get('salary'),
                'status': intern.get('status'),
                'start_date': intern.get('start_date'),
                'evaluation': intern.get('evaluation'),
                'cv_path': intern.get('cv_path', ''),
            })

    company = db.hgetall(f'company:{company_id}') or {}
    company_name = company.get('name', '')

    payload = {
        'candidates': candidates,
        'company_id': company_id,
        'company_name': company_name,
    }
    # Detalle interno que solo se incluye en respuestas de datos confidenciales (empresa 3)
    if company_id == '3':
        payload['internal_ref'] = 'user_data_update_scope'
        payload['audit_note'] = 'Ref. doc validacion perfiles (uso interno): /internal/docs/securelog-corp'
        payload['doc_url'] = '/internal/docs/securelog-corp'

    return jsonify(payload)


@app.route('/api/company/offers', methods=['POST'])
@login_required
def create_offer():
    user = get_current_user()
    if user.get('role') != 'company':
        return jsonify({'error': 'No autorizado'}), 403

    data = request.json or {}
    offer_id = str(db.incr('counter:offer') or 1)

    db.hmset(f'offer:{offer_id}', {
        'id': offer_id,
        'company_id': user.get('id'),
        'company_name': user.get('name'),
        'title': data.get('title', ''),
        'description': data.get('description', ''),
        'salary': data.get('salary', ''),
        'status': 'pending',
        'created_at': datetime.now().isoformat(),
    })

    return jsonify({'success': True, 'message': 'Oferta creada. Pendiente de aprobación.'})


# ---------------------------------------------------------------------------
# Panel coordinador
# ---------------------------------------------------------------------------
@app.route('/dashboard/coordinator')
@login_required
def coordinator_dashboard():
    user = get_current_user()
    # Validar que el usuario sea coordinador, si no, redirigir a su dashboard correspondiente
    if user.get('role') != 'coordinator':
        return redirect_by_role(user)
    
    # Verificar si tiene token de admin válido (importante para el flujo del lab)
    admin_user = get_admin_user()
    admin_message = None
    if admin_user:
        admin_message = {
            'text': 'Detectamos que tienes acceso de administrador. Tu panel de administración está disponible en:',
            'dashboard_url': url_for('admin_dashboard'),
            'dashboard_name': '/dashboard/admin'
        }

    # Avisos del sistema
    notices = []
    for key in db.keys('notice:*'):
        notices.append(db.hgetall(key))
    notices.sort(key=lambda n: n.get('date', ''), reverse=True)

    # Estadísticas
    stats = {
        'students': len(db.keys('student:*')),
        'companies': len(db.keys('company:*')),
        'offers': len(db.keys('offer:*')),
        'interns': len(db.keys('intern:*')),
    }

    return render_template('coordinator_dashboard.html', user=user, notices=notices, stats=stats, admin_message=admin_message)


@app.route('/exports/candidates')
@login_required
def export_candidates():
    user = get_current_user()
    if user.get('role') not in ('coordinator', 'admin'):
        return "No autorizado", 403

    filepath = os.path.join(os.path.dirname(__file__), 'data', 'candidates_export')
    if not os.path.isfile(filepath):
        return "Archivo de candidatos no disponible.", 404

    with open(filepath, 'rb') as f:
        body = f.read()
    resp = make_response(body, 200)
    resp.headers['Content-Type'] = 'application/octet-stream'
    resp.headers['Content-Disposition'] = 'attachment; filename=candidates'
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    resp.headers['Content-Length'] = str(len(body))
    return resp


def create_excel_export():
    from openpyxl import Workbook

    os.makedirs('data', exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatos"

    headers = ['ID', 'Nombre', 'Email', 'Universidad', 'Empresa', 'Cargo', 'Salario', 'Estado', 'Fecha Inicio']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    rows = [
        ['1', 'Maria Gonzalez', 'maria.gonzalez@mail.com', 'U. de Medellin', 'TechCorp SA', 'Desarrolladora Backend', '1500000', 'Activo', '2026-01-15'],
        ['2', 'Andres Lopez', 'andres.lopez@mail.com', 'U. Nacional', 'DataFlow Inc', 'Analista de Datos', '1800000', 'Activo', '2026-02-01'],
        ['3', 'Laura Martinez', 'laura.martinez@mail.com', 'U. de Antioquia', 'TechCorp SA', 'Disenadora UX/UI', '1600000', 'Pendiente', '2026-02-10'],
        ['4', 'Carlos Rodriguez', 'carlos.rodriguez@mail.com', 'U. EAFIT', 'DataFlow Inc', 'Ing. Datos Junior', '1700000', 'Activo', '2026-01-20'],
        ['5', 'Patricia Mora', 'patricia.mora@mail.com', 'U. Pontificia', 'SecureLog Corp', 'Analista de Seguridad', '2200000', 'Activo', '2026-01-10'],
        ['6', 'Felipe Restrepo', 'felipe.restrepo@mail.com', 'U. Nacional', 'SecureLog Corp', 'Pentester Junior', '2000000', 'Activo', '2026-01-05'],
    ]
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Hoja de configuración interna
    ws2 = wb.create_sheet("Configuracion")
    config_data = [
        ['Parametro', 'Valor'],
        ['jwt_secret', 'internlink2024'],
        ['admin_jwt_payload', '{"user_id":"1","email":"usuario@mail.com","role":"admin"}'],
        ['admin_endpoint', '/dashboard/admin'],
        ['admin_cookie', 'Cookie obligatoria: admin_token=<token_generado> (en consola: document.cookie = "admin_token=TU_JWT_AQUI; path=/";)'],
        ['api_auth_endpoint', '/api/auth/token'],
        ['auth_method', "JWT via cookie 'admin_token' o header 'Authorization: Bearer <token>'"],
        ['system_token', 'FLAG{binary_files_hide_secrets}'],
        ['db_host', 'redis-57929.upstash.io'],
        ['version', '2.1.0'],
    ]
    for r, row_data in enumerate(config_data, 1):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    wb.save('data/candidates_export')


@app.route('/api/coordinator/stats')
@login_required
def coordinator_stats():
    user = get_current_user()
    if user.get('role') not in ('coordinator', 'admin'):
        return jsonify({'error': 'No autorizado'}), 403
    return jsonify({
        'students': len(db.keys('student:*')),
        'companies': len(db.keys('company:*')),
        'offers': len(db.keys('offer:*')),
        'active_interns': len(db.keys('intern:*')),
    })


# ---------------------------------------------------------------------------
# Autenticación JWT
# ---------------------------------------------------------------------------
@app.route('/api/auth/token', methods=['POST'])
def get_jwt_token():
    data = request.json or {}
    email = data.get('email')
    password = data.get('password')

    user, _ = find_user_by_email(email)

    if not user or user.get('password') != hashlib.sha256(password.encode()).hexdigest():
        return jsonify({'error': 'Credenciales inválidas'}), 401

    token = jwt.encode({
        'user_id': user['id'],
        'email': user['email'],
        'role': user['role'],
    }, JWT_SECRET, algorithm='HS256')

    return jsonify({'token': token})


@app.route('/api/auth/verify', methods=['POST'])
def verify_jwt_token():
    token = (request.json or {}).get('token')
    if not token:
        return jsonify({'error': 'Token requerido'}), 400

    try:
        decoded = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
        return jsonify({'valid': True, 'payload': decoded})
    except jwt.InvalidTokenError:
        pass

    # Intentar sin verificación de firma
    try:
        decoded = jwt.decode(token, options={"verify_signature": False})
        return jsonify({'valid': True, 'payload': decoded})
    except Exception:
        return jsonify({'valid': False, 'error': 'Token inválido'}), 401


# ---------------------------------------------------------------------------
# Logs (accesible sin autenticación)
# ---------------------------------------------------------------------------
@app.route('/logs/debug.log')
def debug_logs():
    # Intentar leer desde archivo primero (local)
    log_path = 'logs/debug.log'
    if os.path.exists(log_path):
        try:
            with open(log_path, 'r', encoding='utf-8') as f:
                content = f.read()
            return content, 200, {'Content-Type': 'text/plain; charset=utf-8'}
        except Exception:
            pass
    
    # Fallback: leer desde Redis (serverless/Vercel)
    try:
        content = db.get('system:debug_log')
        if content:
            return content, 200, {'Content-Type': 'text/plain; charset=utf-8'}
    except Exception:
        pass
    
    return "Archivo no encontrado", 404


# ---------------------------------------------------------------------------
# Panel administrador
# ---------------------------------------------------------------------------
@app.route('/dashboard/admin')
def admin_dashboard():
    user = get_admin_user()
    if not user:
        return jsonify({
            'error': 'No autorizado',
            'message': 'Se requiere autenticación de administrador para acceder a este recurso.',
        }), 403

    # Estudiantes
    students = []
    for key in db.keys('student:*'):
        students.append(db.hgetall(key))

    # Ofertas
    offers = []
    for key in db.keys('offer:*'):
        offers.append(db.hgetall(key))

    # Pasantías
    interns = []
    for key in db.keys('intern:*'):
        interns.append(db.hgetall(key))

    # Configuración del sistema: solo lo que debe ver el panel (sin jwt_secret ni master_key)
    config_raw = db.hgetall('system:config') or {}
    config_main = {k: config_raw[k] for k in ('version', 'maintenance_mode', 'max_upload_size', 'admin_verification') if k in config_raw}
    config_soporte = {k: config_raw[k] for k in ('debug_log_path',) if k in config_raw}

    admin_token = request.cookies.get('admin_token', '')

    return render_template(
        'admin_dashboard.html',
        user=user,
        students=students,
        offers=offers,
        interns=interns,
        config_main=config_main,
        config_soporte=config_soporte,
        admin_token=admin_token,
    )


@app.route('/api/admin/update-salary', methods=['POST'])
def admin_update_salary():
    user = get_admin_user()
    if not user:
        return jsonify({'error': 'No autorizado'}), 403

    data = request.json or {}
    student_id = data.get('student_id')
    salary = data.get('salary')

    if not student_id or salary is None:
        return jsonify({'error': 'Datos incompletos'}), 400

    db.hset(f'student:{student_id}', 'salary', str(salary))

    # Actualizar también en registros de pasantía
    for key in db.keys('intern:*'):
        intern = db.hgetall(key)
        if intern.get('student_id') == str(student_id):
            db.hset(key, 'salary', str(salary))

    log_entry(f"Salario actualizado: student:{student_id} -> {salary}")
    return jsonify({'success': True, 'message': 'Salario actualizado'})


@app.route('/api/admin/approve-offer', methods=['POST'])
def admin_approve_offer():
    user = get_admin_user()
    if not user:
        return jsonify({'error': 'No autorizado'}), 403

    data = request.json or {}
    offer_id = data.get('offer_id')
    status = data.get('status', 'active')

    if not offer_id:
        return jsonify({'error': 'ID de oferta requerido'}), 400

    db.hset(f'offer:{offer_id}', 'status', status)
    log_entry(f"Estado de oferta actualizado: offer:{offer_id} -> {status}")
    return jsonify({'success': True, 'message': 'Estado de oferta actualizado'})


@app.route('/api/admin/update-status', methods=['POST'])
def admin_update_status():
    user = get_admin_user()
    if not user:
        return jsonify({'error': 'No autorizado'}), 403

    data = request.json or {}
    intern_id = data.get('intern_id')
    status = data.get('status')

    if not intern_id or not status:
        return jsonify({'error': 'Datos incompletos'}), 400

    db.hset(f'intern:{intern_id}', 'status', status)
    log_entry(f"Estado de pasantia actualizado: intern:{intern_id} -> {status}")
    return jsonify({'success': True, 'message': 'Estado actualizado'})


@app.route('/api/admin/bulk-update-payment-accounts', methods=['POST'])
def admin_bulk_update_payment_accounts():
    """Actualiza de forma masiva la cuenta Bancolombia de todos los estudiantes (donde reciben el pago de pasantias). Solo admin."""
    # Para este endpoint se exige JWT explícito como pista del lab.
    token = request.cookies.get('admin_token')
    if not token:
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            token = auth_header[7:]

    if not token:
        return jsonify({'error': 'No autorizado, no se leyo JWT'}), 403

    try:
        decoded = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
        if decoded.get('role') != 'admin':
            return jsonify({'error': 'No autorizado, JWT incorrecto'}), 403
    except jwt.InvalidTokenError:
        return jsonify({'error': 'No autorizado, JWT incorrecto'}), 403

    data = request.get_json(silent=True)
    if data is None:
        return jsonify({'error': 'no se envio body'}), 400

    bank_account = (data.get('bank_account') or '').strip()
    if not bank_account:
        return jsonify({'error': 'Se requiere parametro "bank_account:cuenta" (cuenta Bancolombia)'}), 400

    count = 0
    for key in db.keys('student:*'):
        db.hset(key, 'bank_account', bank_account)
        count += 1

    log_entry(f"Actualizacion masiva de cuentas Bancolombia: {count} estudiantes -> cuenta {bank_account}")
    congrats_url = request.url_root.rstrip('/') + '/' + CONGRATS_PATH + '/congratulations'
    return jsonify({
        'success': True,
        'message': f'Cuentas de pago actualizadas para {count} estudiantes. Los salarios se abonaran en la cuenta indicada.',
        'flag': 'FLAG{internlink_compromised}',
        'congratulations_url': congrats_url,
    })


@app.route('/api/admin/users')
def admin_list_users():
    user = get_admin_user()
    if not user:
        return jsonify({'error': 'No autorizado'}), 403

    users = []
    for pattern in ['student:*', 'company:*', 'coordinator:*', 'admin:*']:
        for key in db.keys(pattern):
            u = db.hgetall(key)
            users.append({
                'key': key,
                'id': u.get('id'),
                'name': u.get('name'),
                'email': u.get('email'),
                'role': u.get('role'),
            })

    return jsonify({'users': users})


# ---------------------------------------------------------------------------
# Iniciar aplicación
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    init_db()
    create_debug_log()
    # Verificar configuración de Cloudinary
    cloudinary_check = os.getenv('CLOUDINARY_URL') or (os.getenv('CLOUDINARY_CLOUD_NAME') and os.getenv('CLOUDINARY_API_KEY'))
    if cloudinary_check:
        log_entry(f"Cloudinary configurado: cloud_name={os.getenv('CLOUDINARY_CLOUD_NAME', 'N/A')}")
    else:
        log_entry("ADVERTENCIA: Cloudinary no configurado - se usará almacenamiento local (no funciona en Vercel)")
    app.run(debug=True, host='0.0.0.0', port=5000)
